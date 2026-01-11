import logging
import os
import pandas as pd
import asyncio
import aiosqlite
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.enums import ChatMemberStatus
from openpyxl.styles import Font, Alignment, PatternFill

from config import ADMIN_IDS, COURSES_CONFIG
from app.keyboards.admin_kb import admin_panel_kb, stats_kb
from app.database.requests import get_all_users_without_admin, search_users, purchase_course, chek_user, get_all_courses_admin, update_course_param, get_course_by_id, get_all_users

router = Router()
logger = logging.getLogger('bot_actions')
DB_NAME = 'Users.db'

class AdminGiveCourse(StatesGroup):
    waiting_for_search = State()
    selecting_user = State()
    selecting_course = State()
    
class EditCourse(StatesGroup):
    waiting_for_name = State()
    waiting_for_price = State()
    waiting_for_id = State()
    # Для добавления нового
    new_c_name = State()
    new_c_price = State()
    new_c_id = State()
    
class EditCourse(StatesGroup):
    waiting_for_new_name = State()
    waiting_for_new_price = State()

@router.message(F.text == "⚙️ Админка")
async def open_admin_panel(message: Message):
    if message.from_user.id in ADMIN_IDS:
        await message.answer("🛠 Панель администратора:", reply_markup=admin_panel_kb)
    else:
        await message.answer("❌ Нет доступа.")

@router.callback_query(F.data == "back_to_admin")
async def back_to_admin(callback: CallbackQuery):
    await callback.message.edit_text("🛠 Панель администратора:", reply_markup=admin_panel_kb)

# --- СТАТИСТИКА ---
@router.callback_query(F.data == "admin_stats")
async def show_stats(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return
    
    users = await get_all_users_without_admin()
    total_users = len(users)
    total_revenue = sum(user[5] for user in users)
    
    c1_cnt = sum(1 for u in users if u[2])
    c2_cnt = sum(1 for u in users if u[3])
    c3_cnt = sum(1 for u in users if u[4])
    
    text = (
        f"📊 **Расширенная статистика**\n\n"
        f"👥 Пользователей: {total_users}\n"
        f"💰 Выручка: {total_revenue} сом\n\n"
        f"📚 **Продажи:**\n"
        f"┣ {COURSES_CONFIG['course_1']['name']}: {c1_cnt}\n"
        f"┣ {COURSES_CONFIG['course_2']['name']}: {c2_cnt}\n"
        f"┗ {COURSES_CONFIG['course_3']['name']}: {c3_cnt}"
    )
    
    try:
        await callback.message.edit_text(text, reply_markup=stats_kb, parse_mode="Markdown")
    except TelegramBadRequest:
        await callback.answer("Данные актуальны")

# --- EXCEL ---
@router.callback_query(F.data == "export_excel")
async def export_excel(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return

    users_data = await get_all_users()
    columns = ['Имя', 'ID', 'Курс 1', 'Курс 2', 'Курс 3', 'Доход']
    df = pd.DataFrame(users_data, columns=columns)

    for c in ['Курс 1', 'Курс 2', 'Курс 3']:
        df[c] = df[c].apply(lambda x: "✅" if x else "❌")

    total_rev = df['Доход'].sum()
    df.loc[len(df)] = ['ИТОГО', '', '', '', '', total_rev]

    file_path = "report.xlsx"
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Отчет')
        ws = writer.sheets['Отчет']
        
        # Стилизация (упрощено для примера)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 30

    await callback.bot.send_document(
        chat_id=callback.from_user.id,
        document=FSInputFile(file_path),
        caption=f"💰 Общая выручка: **{total_rev} сом**",
        parse_mode="Markdown"
    )
    os.remove(file_path)
    await callback.answer()

# --- ВЫДАЧА КУРСА ---
@router.callback_query(F.data == "admin_give_course")
async def start_give(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("🔍 Введите имя или username для поиска:")
    await state.set_state(AdminGiveCourse.waiting_for_search)
    await callback.answer()

@router.message(AdminGiveCourse.waiting_for_search)
async def process_search(message: Message, state: FSMContext):
    users = await search_users(message.text)
    if not users:
        return await message.answer("❌ Не найдено. Попробуйте снова:")
    
    buttons = []
    for uid, name, uname in users:
        label = f"{name} (@{uname})" if uname else name
        buttons.append([InlineKeyboardButton(text=label, callback_data=f"give_to_{uid}")])
    
    await message.answer("👥 Выберите пользователя:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await state.set_state(AdminGiveCourse.selecting_user)

@router.callback_query(F.data.startswith("give_to_"))
async def select_course_give(callback: CallbackQuery, state: FSMContext):
    uid = int(callback.data.replace("give_to_", ""))
    await state.update_data(target_uid=uid)
    
    user = await chek_user(uid)
    # Маппинг для проверки наличия (course_1 -> user[3])
    owned = {"course_1": user[3], "course_2": user[4], "course_3": user[5]}
    
    buttons = []
    for key, data in COURSES_CONFIG.items():
        icon = "✅" if owned[key] else "🎁"
        buttons.append([InlineKeyboardButton(text=f"{icon} {data['name']}", callback_data=f"confirm_give_{key}")])
        
    await callback.message.edit_text(f"📚 Что выдать юзеру {uid}?", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await state.set_state(AdminGiveCourse.selecting_course)

@router.callback_query(F.data.startswith("confirm_give_"))
async def confirm_give(callback: CallbackQuery, state: FSMContext):
    course_key = callback.data.replace("confirm_give_", "")
    data = await state.get_data()
    uid = data['target_uid']
    
    await purchase_course(uid, course_key, 0)
    logger.info(f"MANUAL GIVE: Admin {callback.from_user.id} -> User {uid} -> {course_key}")
    
    await callback.bot.send_message(uid, f"🎁 Админ открыл доступ: **{COURSES_CONFIG[course_key]['name']}**")
    await callback.message.edit_text("✅ Выдано успешно!")
    await state.clear()
    await callback.answer()

# --- ПРОВЕРКА ПОДПИСОК ---
@router.callback_query(F.data == "check_membership")
async def check_members(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return
    await callback.message.answer("⏳ Проверяю участников...")
    
    users = await get_all_users()
    idx_map = {"course_1": 2, "course_2": 3, "course_3": 4}
    
    report = []
    
    for user in users:
        uid, full_name = user[1], user[0]
        
        for key, conf in COURSES_CONFIG.items():
            if user[idx_map[key]] and conf['channel_id']: # Если куплено и есть канал
                try:
                    m = await callback.bot.get_chat_member(conf['channel_id'], uid)
                    if m.status in [ChatMemberStatus.LEFT, ChatMemberStatus.KICKED]:
                        link_obj = await callback.bot.create_chat_invite_link(conf['channel_id'], member_limit=1)
                        msg = f"⚠️ Вы купили **{conf['name']}**, но вышли из канала.\nВернуться: {link_obj.invite_link}"
                        
                        try:
                            await callback.bot.send_message(uid, msg)
                            report.append(f"{full_name} -> {conf['name']} (Уведомлен)")
                        except TelegramForbiddenError:
                            report.append(f"{full_name} -> {conf['name']} (Блок бота)")
                except Exception as e:
                    logger.error(f"Check error {uid}: {e}")
        
        await asyncio.sleep(0.05)
        
    res = f"📢 Итог:\n" + ("\n".join(report) if report else "Все на месте!")
    await callback.message.answer(res)
    await callback.answer()
    
@router.callback_query(F.data == "manage_courses")
async def admin_courses_list(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return

    courses = await get_all_courses_admin()
    builder = InlineKeyboardBuilder()
    
    # courses возвращает кортежи: (course_id, name, price, channel_id, is_active)
    for c_id, name, price, ch_id, is_active in courses:
        # Если is_active == 1 (True), ставим галочку, иначе крестик
        status_icon = "✅" if is_active else "❌"
        # Текст кнопки: "✅ Название | 5000 сом"
        btn_text = f"{status_icon} {name} | {price}"
        builder.row(InlineKeyboardButton(text=btn_text, callback_data=f"manage_c_{c_id}"))
    
    builder.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_admin"))
    
    await callback.message.edit_text(
        "<b>📚 Управление курсами</b>\nВыберите курс для настройки:", 
        reply_markup=builder.as_markup(),
        parse_mode="HTML"
    )

# Меню настроек конкретного курса
@router.callback_query(F.data.startswith("manage_c_"))
async def course_settings(callback: CallbackQuery):
    c_id = callback.data.replace("manage_c_", "")
    
    # Получаем актуальные данные из БД
    course = await get_course_by_id(c_id)
    # course: (course_id, name, price, channel_id, is_active)
    
    name = course[1]
    price = course[2]
    is_active = course[4]

    # Формируем текст кнопки переключения видимости
    # Если сейчас активно (1), предлагаем скрыть (0). И наоборот.
    toggle_text = "🙈 Скрыть курс" if is_active else "👁 Показать курс"
    toggle_callback = f"toggle_{c_id}_{0 if is_active else 1}"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 Изменить название", callback_data=f"editname_{c_id}")],
        [InlineKeyboardButton(text="💰 Изменить цену", callback_data=f"editprice_{c_id}")],
        [InlineKeyboardButton(text=toggle_text, callback_data=toggle_callback)],
        [InlineKeyboardButton(text="⬅️ Назад к списку", callback_data="manage_courses")]
    ])
    
    # Используем HTML, чтобы избежать ошибок с символами
    text = (
        f"⚙️ Настройки курса: <b>{name}</b>\n"
        f"🆔 ID: <code>{c_id}</code>\n"
        f"💸 Текущая цена: <b>{price}</b>\n"
        f"👀 Статус: <b>{'Виден всем' if is_active else 'Скрыт'}</b>"
    )
    
    await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")

# --- 1. ЛОГИКА ИЗМЕНЕНИЯ ИМЕНИ ---

@router.callback_query(F.data.startswith("editname_"))
async def start_edit_name(callback: CallbackQuery, state: FSMContext):
    c_id = callback.data.replace("editname_", "")
    # Запоминаем, какой курс редактируем
    await state.update_data(target_course_id=c_id)
    
    await callback.message.edit_text(
        f"✍️ Введите новое <b>название</b> для курса <code>{c_id}</code>:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Отмена", callback_data=f"manage_c_{c_id}")]
        ])
    )
    await state.set_state(EditCourse.waiting_for_new_name)

@router.message(EditCourse.waiting_for_new_name)
async def process_new_name(message: Message, state: FSMContext):
    data = await state.get_data()
    c_id = data.get("target_course_id")
    new_name = message.text

    # Обновляем в БД
    await update_course_param(c_id, "name", new_name)
    
    await message.answer(f"✅ Название курса обновлено на: <b>{new_name}</b>", parse_mode="HTML")
    
    # Сбрасываем состояние
    await state.clear()
    
    # Показываем меню этого курса снова (эмулируем нажатие кнопки)
    # Нам нужно снова сформировать меню, поэтому вызываем логику course_settings вручную или просим юзера нажать кнопку
    # Проще отправить сообщение с кнопкой "Вернуться"
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⚙️ Вернуться к настройкам", callback_data=f"manage_c_{c_id}")]
    ])
    await message.answer("Что делаем дальше?", reply_markup=kb)

# --- 2. ЛОГИКА ИЗМЕНЕНИЯ ЦЕНЫ ---

@router.callback_query(F.data.startswith("editprice_"))
async def start_edit_price(callback: CallbackQuery, state: FSMContext):
    c_id = callback.data.replace("editprice_", "")
    await state.update_data(target_course_id=c_id)
    
    await callback.message.edit_text(
        f"💰 Введите новую <b>цену</b> (только цифры) для курса <code>{c_id}</code>:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Отмена", callback_data=f"manage_c_{c_id}")]
        ])
    )
    await state.set_state(EditCourse.waiting_for_new_price)

@router.message(EditCourse.waiting_for_new_price)
async def process_new_price(message: Message, state: FSMContext):
    if not message.text.isdigit():
        return await message.answer("⚠️ Ошибка: цена должна состоять только из цифр. Попробуйте снова.")
    
    data = await state.get_data()
    c_id = data.get("target_course_id")
    new_price = int(message.text)

    # Обновляем в БД
    await update_course_param(c_id, "price", new_price)
    
    await message.answer(f"✅ Цена обновлена: <b>{new_price}</b>", parse_mode="HTML")
    await state.clear()
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⚙️ Вернуться к настройкам", callback_data=f"manage_c_{c_id}")]
    ])
    await message.answer("Что делаем дальше?", reply_markup=kb)

# --- 3. ЛОГИКА СКРЫТИЯ/ПОКАЗА ---

@router.callback_query(F.data.startswith("toggle_"))
async def toggle_visibility(callback: CallbackQuery):
    # data имеет вид: toggle_{course_id}_{new_status}
    parts = callback.data.split("_")
    
    # Последний элемент - статус (0 или 1)
    new_status = int(parts[-1])
    # Всё, что между "toggle" и статусом - это ID курса
    c_id = "_".join(parts[1:-1]) 
    
    # 1. Обновляем статус в БД
    await update_course_param(c_id, "is_active", new_status)
    
    # 2. Показываем уведомление
    status_text = "скрыт 🙈" if new_status == 0 else "теперь виден всем 👁"
    await callback.answer(f"Курс {status_text}")
    
    # 3. ЧТОБЫ ОБНОВИТЬ МЕНЮ БЕЗ ОШИБКИ:
    # Заново получаем данные курса и рисуем клавиатуру прямо здесь
    course = await get_course_by_id(c_id)
    # course: (course_id, name, price, channel_id, is_active)
    
    name = course[1]
    price = course[2]
    is_active = course[4] # Это уже обновленный статус

    toggle_text = "🙈 Скрыть курс" if is_active else "👁 Показать курс"
    toggle_callback = f"toggle_{c_id}_{0 if is_active else 1}"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 Изменить название", callback_data=f"editname_{c_id}")],
        [InlineKeyboardButton(text="💰 Изменить цену", callback_data=f"editprice_{c_id}")],
        [InlineKeyboardButton(text=toggle_text, callback_data=toggle_callback)],
        [InlineKeyboardButton(text="⬅️ Назад к списку", callback_data="manage_courses")]
    ])
    
    text = (
        f"⚙️ Настройки курса: <b>{name}</b>\n"
        f"🆔 ID: <code>{c_id}</code>\n"
        f"💸 Текущая цена: <b>{price}</b>\n"
        f"👀 Статус: <b>{'Виден всем' if is_active else 'Скрыт'}</b>"
    )
    
    # Редактируем сообщение
    await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")