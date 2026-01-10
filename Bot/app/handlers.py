import os
import pandas as pd
import logging
import asyncio
from aiogram import F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import FSInputFile, Message, CallbackQuery
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.enums import ChatMemberStatus
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.types import ChatJoinRequest

# Импорты из ваших файлов проекта
from config import ADMIN_IDS, COURSES_CONFIG # Изменено на ADMIN_IDS
from app.keyboard import keyboard_start, get_buy_keyboard, keyboard_decision, keyboard_admin, get_buy_keyboard
from app.methods import add_user, update_user_name, chek_user, purchase_course, get_all_users, search_users, chek_user_funk
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger('bot_actions')

r = Router()

# ОГРАНИЧЕНИЕ: Только личные сообщения
# r.message.filter(F.chat.type == ChatType.PRIVATE)
# r.callback_query.filter(F.chat.type == ChatType.PRIVATE)

if not os.path.exists("Assets"):
    os.makedirs("Assets")
if not os.path.exists("payments"):
    os.makedirs("payments")

photo_file = FSInputFile("Assets/payment.jpg")

class Registration(StatesGroup):
    waiting_for_name = State()

class BuyCourse(StatesGroup):
    waiting_for_course_image = State()
    course = State()
    
class AdminGiveCourse(StatesGroup):
    waiting_for_search = State()   # Ожидание ввода имени для поиска
    selecting_user = State()       # Выбор юзера из списка
    selecting_course = State()     # Выбор курса для выдачи

# --- ПОЛЬЗОВАТЕЛЬСКИЕ КОМАНДЫ ---

@r.message(Command("start"))
async def send_welcome(message: Message, state: FSMContext):
    user = await chek_user(message.from_user.id)
    if not user:
        await add_user(message.from_user.id, message.from_user.username, message.chat.id)
        await message.answer("👋 Здравствуйте! Пожалуйста, введите вашу **Фамилию и Имя**:")
        await state.set_state(Registration.waiting_for_name)
    else:
        await message.answer(f"С возвращением, {user[2] if user[2] else 'друг'}!", reply_markup=keyboard_start)
        
@r.message(Registration.waiting_for_name)
async def process_name(message: Message, state: FSMContext):
    logger.info(f"REGISTRATION COMPLETED: User {message.from_user.id} set name to '{message.text}'")
    await update_user_name(message.from_user.id, message.text)
    await state.clear()
    await message.answer(f"Приятно познакомиться, {message.text}! Теперь вам доступно меню.", reply_markup=keyboard_start)

# В файле handlers.py
@r.message(Command("buy"))
async def buy_course(message: Message, state: FSMContext):
    await chek_user_funk(message, message.from_user.id, state)
    user = await chek_user(message.from_user.id)
    await message.answer(
        "🛒 **Выберите курс для покупки:**\n\n"
        "Ваши курсы отмечены галочкой ✅",
        reply_markup=get_buy_keyboard(user),
        parse_mode="Markdown"
    )

# Добавим обработку нажатия на уже купленный курс
@r.callback_query(F.data == "already_owned")
async def owned_info(callback_query: CallbackQuery):
    await callback_query.answer("🌟 Этот курс уже доступен вам!", show_alert=True)

# --- АДМИН ПАНЕЛЬ ---

@r.message(F.text == "⚙️ Админка")
async def admin_panel(message: Message):

    if message.from_user.id in ADMIN_IDS:
        await message.answer("🛠 Панель администратора:", reply_markup=keyboard_admin)
    else:
        await message.answer("❌ У вас нет прав доступа.")

@r.callback_query(F.data == "export_excel")
async def export_to_excel(callback_query: CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS: return

    users_data = await get_all_users()
    

    columns = ['Имя', 'ID', 'Курс 1', 'Курс 2', 'Курс 3', 'Доход с юзера']
    df = pd.DataFrame(users_data, columns=columns)


    for c in ['Курс 1', 'Курс 2', 'Курс 3']:
        df[c] = df[c].apply(lambda x: "✅" if x else "❌")


    total_revenue = df['Доход с юзера'].sum()
    

    total_row = pd.DataFrame([['ИТОГО', '', '', '', '', total_revenue]], columns=columns)
    df = pd.concat([df, total_row], ignore_index=True)

    file_path = "report.xlsx"
    

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Отчет')
        

        workbook = writer.book
        worksheet = writer.sheets['Отчет']


        worksheet.column_dimensions['A'].width = 40  # Имя (Широкое!)
        worksheet.column_dimensions['B'].width = 15  # ID
        worksheet.column_dimensions['C'].width = 12  # Курс 1
        worksheet.column_dimensions['D'].width = 12  # Курс 2
        worksheet.column_dimensions['E'].width = 12  # Курс 3
        worksheet.column_dimensions['F'].width = 20  # Доход

        # Определяем стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Проходимся по всем строкам данных
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter in ['A']:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

        last_row = len(df) + 1
        for cell in worksheet[last_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    await callback_query.bot.send_document(
        chat_id=callback_query.from_user.id,
        document=FSInputFile(file_path),
        caption=f"📊 Финансовый отчет\n💰 Общая выручка: **{total_revenue} сом**",
        parse_mode="Markdown"
    )
    os.remove(file_path)
    await callback_query.answer()

# --- ПРОЦЕСС ПОКУПКИ ---

@r.callback_query(F.data.startswith("buy_"))
async def process_buy_callback(callback_query: CallbackQuery, state: FSMContext):
    course_key = callback_query.data.replace("buy_", "")
    course_data = COURSES_CONFIG.get(course_key)
    
    await callback_query.message.answer_photo(
        photo=photo_file,
        caption=f"📍 Курс: **{course_data['name']}**\n💰 Цена: **{course_data['price']}**\n\nОтправьте скриншот оплаты.",
        parse_mode="Markdown"
    )
    await state.set_state(BuyCourse.waiting_for_course_image)
    await state.update_data(course=course_key)
    await callback_query.answer()

@r.message(BuyCourse.waiting_for_course_image, F.photo)
async def process_course_image(message: Message, state: FSMContext):
    user_data = await state.get_data()
    course_key = user_data.get('course')
    course_name = COURSES_CONFIG[course_key]['name']
    
    logger.info(f"PURCHASE REQUEST: User {message.from_user.id} uploaded receipt for '{course_name}'")
    
    await message.answer("📥 Скриншот получен. Ждите одобрения!")
    
    photo = message.photo[-1]
    file_path = f"payments/{message.from_user.id}_{course_key}.jpg"
    await message.bot.download(file=photo.file_id, destination=file_path)
    
    # Отправка уведомления ВСЕМ админам
    admin_photo = FSInputFile(file_path)
    for admin_id in ADMIN_IDS:
        try:
            await message.bot.send_photo(
                chat_id=admin_id,
                photo=admin_photo,
                caption=f"🔔 **Новая заявка!**\n👤 Юзер: @{message.from_user.username}\n🆔 ID: {message.from_user.id}\n📚 Курс: {course_name}",
                reply_markup=keyboard_decision(message.from_user.id, course_key),
                parse_mode="Markdown"
            )
        except Exception as e:
            print(f"Не удалось отправить уведомление админу {admin_id}: {e}")
    
    await state.clear()

# --- ОДОБРЕНИЕ / ОТКЛОНЕНИЕ ---

@r.callback_query(F.data.startswith("accept-"))
async def decision_yes(callback_query: CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS:
        return await callback_query.answer("Нет прав")

    _, user_id, course_key = callback_query.data.split("-")
    user_id = int(user_id)
    course_data = COURSES_CONFIG.get(course_key)
    price_str = COURSES_CONFIG[course_key]['price']
    price_int = int(''.join(filter(str.isdigit, price_str)))
    
    try:
        await purchase_course(int(user_id), course_key, price_int)
        
        # ЛОГ: Одобрение покупки
        logger.info(f"PURCHASE APPROVED: Admin {callback_query.from_user.username} approved '{course_data['name']}' for User {user_id}")

        invite_text = ""
        if course_data.get("channel_id"):
            try:
                link = await callback_query.bot.create_chat_invite_link(course_data["channel_id"], member_limit=1)
                invite_text = f"\n\n🔗 Ссылка: {link.invite_link}"
            except Exception as e:
                logger.error(f"LINK CREATION FAILED: {e}")
                invite_text = "\n\n⚠️ Ссылка не создана, свяжитесь с админом."

        await callback_query.bot.send_message(
            chat_id=user_id, 
            text=f"✅ Покупка **{course_data['name']}** одобрена!{invite_text}",
            parse_mode="Markdown"
        )
        await callback_query.message.edit_caption(
            caption=callback_query.message.caption + f"\n\n🟢 Одобрено админом @{callback_query.from_user.username}"
        )
    except Exception as e:
        logger.error(f"ERROR APPROVING: {e}")
        await callback_query.message.answer(f"Ошибка: {e}")
    
    await callback_query.answer()
    await callback_query.message.delete()

@r.callback_query(F.data.startswith("decline-"))
async def decision_no(callback_query: CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS:
        return await callback_query.answer("Нет прав")

    _, user_id, course_key = callback_query.data.split("-")
    user_id = int(user_id)
    
    # ЛОГ: Отклонение
    logger.info(f"PURCHASE DECLINED: Admin {callback_query.from_user.username} declined course for User {user_id}")
    
    await callback_query.bot.send_message(user_id, "❌ Ваша оплата отклонена.")
    await callback_query.message.edit_caption(
        caption=callback_query.message.caption + f"\n\n🔴 Отклонено админом @{callback_query.from_user.username}"
    )
    
    await callback_query.message.delete()
    await callback_query.answer()
    
# В файле handlers.py
@r.callback_query(F.data == "admin_stats")
async def admin_stats(callback_query: CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS: return
    
    users = await get_all_users()
    total_users = len(users)
    total_revenue = sum(user[5] for user in users)
    
    # Считаем количество покупок по каждому курсу
    # Индексы курсов в выборке get_all_users: course_1 (2), course_2 (3), course_3 (4)
    c1_count = sum(1 for user in users if user[2])
    c2_count = sum(1 for user in users if user[3])
    c3_count = sum(1 for user in users if user[4])
    
    stats_text = (
        f"📊 **Расширенная статистика**\n\n"
        f"👥 Всего пользователей: {total_users}\n"
        f"💰 Общая выручка: {total_revenue} сом\n\n"
        f"📚 **Продажи по курсам:**\n"
        f"┣ {COURSES_CONFIG['course_1']['name']}: {c1_count} шт.\n"
        f"┣ {COURSES_CONFIG['course_2']['name']}: {c2_count} шт.\n"
        f"┗ {COURSES_CONFIG['course_3']['name']}: {c3_count} шт."
    )
    
    # Кнопка для возврата или обновления
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_stats")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_admin")]
    ])
    
    # await callback_query.message.edit_text(stats_text, reply_markup=kb, parse_mode="Markdown")
    try:
        await callback_query.message.edit_text(stats_text, reply_markup=kb, parse_mode="Markdown")
        # await callback_query.answer("Данные обновлены")
    except TelegramBadRequest as e:
        if "message is not modified" in str(e):
            # Если данные не изменились, просто убираем "часики" на кнопке
            await callback_query.answer("Новых данных пока нет")
        else:
            # Если возникла другая ошибка — пробрасываем её дальше
            raise e
    
# --- ВЫДАЧА КУРСА ВРУЧНУЮ (handlers.py) ---

@r.callback_query(F.data == "admin_give_course")
async def start_give_course(callback_query: CallbackQuery, state: FSMContext):
    if callback_query.from_user.id not in ADMIN_IDS: return
    await callback_query.message.answer("🔍 Введите имя или @username пользователя для поиска:")
    await state.set_state(AdminGiveCourse.waiting_for_search)
    await callback_query.answer()

@r.message(AdminGiveCourse.waiting_for_search)
async def process_user_search(message: Message, state: FSMContext):
    users = await search_users(message.text)
    if not users:
        return await message.answer("❌ Пользователь не найден. Попробуйте еще раз:")
    
    # Создаем кнопки с найденными пользователями
    buttons = []
    for u_id, name, username in users:
        label = f"{name} (@{username})" if username else name
        buttons.append([InlineKeyboardButton(text=label, callback_data=f"give_to_{u_id}")])
    
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await message.answer("👥 Выберите пользователя из списка:", reply_markup=kb)
    await state.set_state(AdminGiveCourse.selecting_user)

@r.callback_query(F.data.startswith("give_to_"))
async def select_course_for_user(callback_query: CallbackQuery, state: FSMContext):
    user_id = int(callback_query.data.replace("give_to_", ""))
    await state.update_data(target_user_id=user_id)
    
    user_data = await chek_user(user_id)
    owned_status = {"course_1": user_data[3], "course_2": user_data[4], "course_3": user_data[5]}

    buttons = []
    for key, data in COURSES_CONFIG.items():
        status_icon = "✅" if owned_status.get(key) else "❌"
        buttons.append([InlineKeyboardButton(text=f"{status_icon} {data['name']}", callback_data=f"confirm_give_{key}")])
    
    await callback_query.message.edit_text(f"📚 Выбор курса для юзера ID: {user_id}", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await state.set_state(AdminGiveCourse.selecting_course)

@r.callback_query(F.data.startswith("confirm_give_"))
async def finalize_give_course(callback_query: CallbackQuery, state: FSMContext):
    course_key = callback_query.data.replace("confirm_give_", "")
    data = await state.get_data()
    user_id = int(data['target_user_id'])
    
    await purchase_course(user_id, course_key, 0)
    
    # ЛОГ: Ручная выдача
    logger.info(f"MANUAL GRANT: Admin {callback_query.from_user.username} gave '{course_key}' to User {user_id}")
    
    await callback_query.bot.send_message(user_id, f"🎁 Администратор открыл вам доступ к курсу: **{COURSES_CONFIG[course_key]['name']}**")
    await callback_query.message.edit_text(f"✅ Курс успешно выдан пользователю ID: {user_id}")
    await state.clear()
    await callback_query.answer()
    
@r.callback_query(F.data == "back_to_admin")
async def back_to_admin(callback_query: CallbackQuery):
    if callback_query.from_user.id in ADMIN_IDS:
        await callback_query.message.edit_text("🛠 Панель администратора:", reply_markup=keyboard_admin)
        
@r.chat_join_request()
async def handle_join_request(update: ChatJoinRequest):
    user_id = update.from_user.id
    chat_id = update.chat.id
    
    target_course_key = None
    for key, data in COURSES_CONFIG.items():
        if data.get("channel_id") == chat_id:
            target_course_key = key
            break

    if not target_course_key:
        return

    user = await chek_user(user_id)
    mapping = {"course_1": 3, "course_2": 4, "course_3": 5}
    
    is_bought = False
    if user:
        is_bought = user[mapping[target_course_key]]
        if user_id in ADMIN_IDS:
            is_bought = True 

    if is_bought:
        await update.approve()
        # ЛОГ: Вступление в группу
        logger.info(f"GROUP JOIN: User {user_id} joined '{COURSES_CONFIG[target_course_key]['name']}' (Auto-approved)")
        
        await update.bot.send_message(
            user_id, 
            f"✅ Ваша заявка в группу **{update.chat.title}** одобрена автоматически!"
        )
    else:
        # ЛОГ: Попытка входа без оплаты
        logger.warning(f"GROUP JOIN DENIED: User {user_id} tried to join '{COURSES_CONFIG[target_course_key]['name']}' without paying.")
        
        await update.decline()
        await update.bot.send_message(
            user_id, 
            f"❌ Доступ в группу **{update.chat.title}** ограничен.\nСначала необходимо приобрести курс.",
            reply_markup=get_buy_keyboard(user)
        )
        
@r.callback_query(F.data == "check_membership")
async def check_membership_status(callback_query: CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS: return
    
    await callback_query.message.answer("⏳ Начинаю проверку участников... Это может занять время.")
    
    # Получаем всех (full_name, user_id, course_1, course_2, course_3, total_spent)
    users = await get_all_users()
    
    # Маппинг ключей конфига на индексы в БД
    # course_1 -> users[i][2], course_2 -> users[i][3], course_3 -> users[i][4]
    course_indices = {"course_1": 2, "course_2": 3, "course_3": 4}
    
    missing_report = []
    count_missing = 0
    
    for user in users:
        user_id = user[1]
        full_name = user[0]
        
        for key, data in COURSES_CONFIG.items():
            # Если курс куплен в БД (значение 1)
            if user[course_indices[key]]:
                channel_id = data.get("channel_id")
                
                if not channel_id:
                    continue # Если у курса нет канала в конфиге
                
                try:
                    member = await callback_query.bot.get_chat_member(chat_id=channel_id, user_id=user_id)
                    
                    # Статусы, когда юзера нет в канале
                    if member.status in [ChatMemberStatus.LEFT, ChatMemberStatus.KICKED, ChatMemberStatus.RESTRICTED]:
                        count_missing += 1
                        
                        # Создаем ссылку
                        link = ""
                        try:
                            # Создаем одноразовую ссылку
                            invite = await callback_query.bot.create_chat_invite_link(channel_id, member_limit=1)
                            link = invite.invite_link
                        except:
                            link = "Обратитесь к администратору."
                        
                        # Отправляем сообщение юзеру
                        msg_text = (
                            f"⚠️ **Внимание!**\n"
                            f"Мы заметили, что вы купили курс **{data['name']}**, но вас нет в канале.\n"
                            f"Пожалуйста, вступите по ссылке: {link}"
                        )
                        
                        try:
                            await callback_query.bot.send_message(user_id, msg_text, parse_mode="Markdown")
                            # ЛОГ: Уведомление об отсутствии
                            logger.info(f"MISSING USER ALERT: Sent alert to User {user_id} for course {key}")
                            missing_report.append(f"👤 {full_name} (ID: {user_id}) -> {data['name']} [Уведомлен]")
                        except TelegramForbiddenError:
                            # Бот заблокирован пользователем
                            logger.warning(f"FAILED ALERT: User {user_id} blocked the bot.")
                            missing_report.append(f"👤 {full_name} (ID: {user_id}) -> {data['name']} [Бот заблокирован]")
                            
                except TelegramBadRequest:
                    # Скорее всего бот не админ в канале или канал неверен
                    logger.error(f"CHECK ERROR: Bot cannot check member in channel {channel_id}")
                except Exception as e:
                    logger.error(f"UNKNOWN ERROR checking user {user_id}: {e}")
        
        # Небольшая пауза, чтобы не словить FloodWait при большом кол-ве юзеров
        await asyncio.sleep(0.05)

    report_text = f"📢 **Результаты проверки:**\n\nНайдено отсутствующих: {count_missing}\n\n"
    if missing_report:
        report_text += "\n".join(missing_report[:50]) # Показываем первые 50, чтобы сообщение не обрезалось
        if len(missing_report) > 50:
            report_text += f"\n... и еще {len(missing_report) - 50}"
    else:
        report_text += "Все пользователи с оплатой находятся в группах! 🎉"
        
    await callback_query.message.answer(report_text)
    await callback_query.answer()